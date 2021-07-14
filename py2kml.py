import openpyxl
import warnings
import sys
import os

colors = {"red":"ed4543", "blue":"177bc9", "green":"56db40",
          "yellow":"ffd21e"}




def file_header(out):
    with open("geojson_begining.txt") as f:
        with open(out, "w") as k:
            for line in f:
                k.write(line)


def file_ending(out):
    with open(out, "rb+") as l:
        l.seek(-1, os.SEEK_END)
        l.truncate()
    with open("geojson_ending.txt") as f:
        with open(out, "a") as k:
            for line in f:
                k.write(line)




def parser(f):
    wb = openpyxl.load_workbook(filename = f)
    ws = wb.active

    output = "output"
    file_header(output+".geojson")


    # default color -> Red
    color = "ed4543"

    for row in ws.iter_rows(min_col=3, max_col=3, min_row=2):
        for cell in row:
            descr = str((cell.value))
            print (descr)
            coord_lat = str(ws['A'+str(cell.row)].value)
            print (coord_lat)
            coord_long = str(ws['B'+str(cell.row)].value)
            print (coord_long)
            color = str(colors[ws['F'+str(cell.row)].value])
            print (color)
            number = str(ws['E'+str(cell.row)].value)
            print (number)
            label = str(ws['D'+str(cell.row)].value)
            print (label)

            with open(output+".geojson", "a") as k:
                k.write("{\"type\":\"Feature\",\"id\":"+number+",\"geometry\":{\"type\":\"Point\",\"coordinates\":["+coord_long+","+coord_lat+"]},\"properties\":{\"description\":\""+descr+"\",\"iconCaption\":\""+label+"\",\"iconContent\":\""+number+"\",\"marker-color\":\""+color+"\"}},")
                k.close()
            print (label+" записана!\n")  

       
    file_ending(output+".geojson")
    print("\n*** Файлик "+output+".geojson записался успешно! ***")


if __name__ == '__main__':
    warnings.filterwarnings("ignore")
    path = "input"
    parser(path+".xlsx")